import streamlit as st
import json
import pandas as pd
import io
import os
import re
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

st.set_page_config(page_title="Simetrik Docs Pro | PeYa", page_icon="📊", layout="wide", initial_sidebar_state="expanded")

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════
C = {
    "red":    "EA0050", "white": "FFFFFF", "grey":  "F5F5F5",
    "dark":   "1C1C1C", "border":"D8D8D8", "blue":  "1565C0",
    "teal":   "00695C", "amber": "E65100", "purple":"4A148C",
    "green":  "1B5E20", "slate": "37474F", "rose":  "880E4F",
}

RT_LABEL = {
    "native":                  "📥 Fuente",
    "source_union":            "🔗 Unión de Fuentes",
    "source_group":            "📊 Agrupación (Group By)",
    "reconciliation":          "⚖️ Conciliación Estándar",
    "advanced_reconciliation": "🔬 Conciliación Avanzada",
    "consolidation":           "🗂️ Consolidación",
    "resource_join":           "🔀 Join de Recursos",
    "cumulative_balance":      "📈 Balance Acumulado",
}

RT_COLOR = {
    "native":                  C["blue"],
    "source_union":            C["teal"],
    "source_group":            C["amber"],
    "reconciliation":          C["red"],
    "advanced_reconciliation": C["purple"],
    "consolidation":           C["slate"],
    "resource_join":           C["green"],
    "cumulative_balance":      C["green"],
}

RT_ORDER = {
    "native": 1, "source_union": 2, "source_group": 3,
    "reconciliation": 4, "advanced_reconciliation": 5,
    "consolidation": 6, "resource_join": 7, "cumulative_balance": 8,
}

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS OPENPYXL
# ══════════════════════════════════════════════════════════════════════════════
def mk_border():
    t = Side(border_style="thin", color=C["border"])
    return Border(left=t, right=t, top=t, bottom=t)

def sc(cell, bg=None, bold=False, color=C["dark"], size=10, ha='left', va='top', wrap=True):
    cell.border = mk_border()
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    cell.font = Font(name='Arial', bold=bold, size=size, color=color)
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

def hdr(cell, text, bg=C["dark"]):
    cell.value = text
    sc(cell, bg=bg, bold=True, color=C["white"], size=10, ha='center', va='center', wrap=False)

def section_title(ws, row, text, bg=C["red"], cols=5):
    ws.merge_cells(f'A{row}:{chr(64+cols)}{row}')
    c = ws.cell(row, 1, text)
    sc(c, bg=bg, bold=True, color=C["white"], size=10, ha='left', va='center', wrap=False)
    ws.row_dimensions[row].height = 20
    return row + 1

def meta_row(ws, row, label, value, cols=5, bg_val=None):
    bg_val = bg_val or C["grey"]
    c_l = ws.cell(row, 1, label)
    sc(c_l, bg=C["slate"], bold=True, color=C["white"], size=9, ha='left', va='center', wrap=False)
    ws.merge_cells(f'B{row}:{chr(64+cols)}{row}')
    c_v = ws.cell(row, 2, str(value) if value is not None else "—")
    sc(c_v, bg=bg_val, size=9, va='center', wrap=True)
    ws.row_dimensions[row].height = 14
    return row + 1

def row_height(n_lines, base=13):
    return max(14, n_lines * base)

# ══════════════════════════════════════════════════════════════════════════════
# PARSERS
# ══════════════════════════════════════════════════════════════════════════════
def build_maps(data):
    res_map = {}
    col_map = {}
    seg_map = {}
    meta_map = {}

    for r in data.get('resources', []):
        eid = r.get('export_id')
        if not eid:
            continue
            
        res_map[eid] = r.get('name', str(eid))

        for c in (r.get('columns') or []):
            cid = c.get('export_id')
            if cid:
                col_map[cid] = c.get('label') or c.get('name') or str(cid)

        sg = r.get('source_group') or {}
        for c in sg.get('columns', []):
            cid = c.get('column_id')
            if cid and cid not in col_map:
                col_map[cid] = f"col_{cid}"
                
        for v in sg.get('values', []):
            cid = v.get('column_id')
            if cid and cid not in col_map:
                col_map[cid] = f"col_{cid}"

        adv = r.get('advanced_reconciliation') or {}
        for rg in adv.get('reconcilable_groups', []):
            for cs in rg.get('columns_selection', []):
                cid = cs.get('column_id')
                if cid and cid not in col_map:
                    col_map[cid] = f"col_{cid}"
            sc2 = rg.get('segmentation_config') or {}
            for m in sc2.get('segmentation_metadata', []):
                meta_id = m.get('export_id')
                if meta_id:
                    meta_map[meta_id] = m.get('value', '?')
            ccid = sc2.get('criteria_column_id')
            if ccid and ccid not in col_map:
                col_map[ccid] = f"col_{ccid}"

        for seg in (r.get('segments') or []):
            rules = []
            for fset in (seg.get('segment_filter_sets') or []):
                for rule in (fset.get('segment_filter_rules') or []):
                    rules.append(rule)
            
            seg_id = seg.get('export_id')
            if seg_id:
                seg_map[seg_id] = {
                    'name':        seg.get('name', ''),
                    'resource':    r.get('name', ''),
                    'resource_id': eid,
                    'rules':       rules,
                }

    return res_map, col_map, seg_map, meta_map

def fmt_filter_rules(rules, col_map):
    lines = []
    for r in rules:
        col_name = col_map.get(r.get('column_id'), f"ID:{r.get('column_id')}")
        cond = r.get('condition', '')
        op = r.get('operator', '')
        val = r.get('value', '')
        lines.append(f"{cond} [{col_name}] {op} {val}".strip())
    return "\n".join(lines) if lines else "Sin filtros adicionales"

def parse_transformation_logic(col, res_map, col_map):
    lines = []
    
    # 1. Búsquedas VLOOKUP
    v = col.get('v_lookup')
    if v:
        vs = v.get('v_lookup_set') or {}
        origin_id = vs.get('origin_source_id')
        origin = res_map.get(origin_id, f"ID:{origin_id}")
        rules = vs.get('rules', [])
        keys = " & ".join(
            "A." + col_map.get(r.get('column_a_id'), '?') +
            " = B." + col_map.get(r.get('column_b_id'), '?')
            for r in rules
        )
        lines.append(f"🔍 BUSCAR V EN: {origin}")
        if keys:
            lines.append(f"🔑 CLAVE MATCH: {keys}")

    # 2. Iterar transformaciones
    transformations = col.get('transformations') or []
    for t in transformations:
        t_type = str(t.get('type', '')).lower()
        
        # Fórmulas
        if t.get('is_parent'):
            q = (t.get('query') or '').strip()
            if q and q.upper() != 'N/A':
                lines.append(f"⚙️ FÓRMULA: {q}")
                
        # Duplicados
        if t_type in ['duplicate', 'row_number'] or 'partition_by' in t:
            dup_label = "Booleano (Flag)" if t_type == 'duplicate' else "Numérico (Índice)"
            lines.append(f"👯 CONTROL DUPLICADOS [{dup_label}]")
            
            partitions = t.get('partition_by') or []
            if partitions:
                part_names = []
                for p in partitions:
                    p_id = p.get('column_id') if isinstance(p, dict) else p
                    part_names.append(col_map.get(p_id, f"ID:{p_id}"))
                lines.append(f"   ├─ Partición: {', '.join(part_names)}")
                
            orders = t.get('order_by') or []
            if orders:
                order_strs = []
                for o in orders:
                    col_id = o.get('column_id')
                    direction = str(o.get('direction', 'ASC')).upper()
                    order_strs.append(f"{col_map.get(col_id, f'ID:{col_id}')} ({direction})")
                lines.append(f"   └─ Orden: {', '.join(order_strs)}")

    return "\n".join(lines) if lines else "Dato directo / heredado"

def parse_std_reconciliation(recon, res_map, col_map, seg_map):
    if not recon:
        return None

    sa_id = recon.get('segment_a_id')
    sb_id = recon.get('segment_b_id')
    a_cfg = recon.get('a_source_settings') or {}
    b_cfg = recon.get('b_source_settings') or {}

    def resolve_side(cfg, seg_id, prefix):
        resource_name = res_map.get(cfg.get('resource_id'), '—')
        seg = seg_map.get(seg_id) or {}
        seg_name = seg.get('name', f"ID:{seg_id}")
        seg_rules = fmt_filter_rules(seg.get('rules', []), col_map)
        return {
            'prefix':        prefix,
            'resource_name': resource_name,
            'group_name':    seg_name,
            'group_filters': seg_rules,
            'is_trigger':    cfg.get('is_trigger', False),
        }

    sides = [
        resolve_side(a_cfg, sa_id, recon.get('segment_a_prefix', 'A')),
        resolve_side(b_cfg, sb_id, recon.get('segment_b_prefix', 'B')),
    ]

    rule_sets = []
    for rs in sorted(recon.get('reconciliation_rule_sets', []), key=lambda x: x.get('position', 99)):
        rules_desc = []
        for rule in (rs.get('reconciliation_rules') or []):
            col_a = col_map.get(rule.get('column_a_id'), f"ID:{rule.get('column_a_id')}")
            col_b = col_map.get(rule.get('column_b_id'), f"ID:{rule.get('column_b_id')}")
            op = rule.get('operator', '=')
            tol = rule.get('tolerance', 0)
            tol_u = rule.get('tolerance_unit') or ''
            tol_s = f"  [tolerancia ±{tol} {tol_u}]" if tol else ""
            rules_desc.append(f"A.{col_a}  {op}  B.{col_b}{tol_s}")
            
        rule_sets.append({
            'pos':   rs.get('position', 0),
            'name':  rs.get('name', ''),
            'rules': rules_desc,
        })

    return {
        'sides':      sides,
        'is_chained': recon.get('is_chained', False),
        'rule_sets':  rule_sets,
    }

def parse_adv_reconciliation(adv, res_map, col_map, seg_map, meta_map):
    if not adv:
        return None

    groups = []
    for rg in (adv.get('reconcilable_groups') or []):
        prefix = rg.get('prefix_side', '?')
        seg_id = rg.get('segment_id')
        seg = seg_map.get(seg_id) or {}
        seg_name = seg.get('name', f"ID:{seg_id}")
        resource_name = seg.get('resource', res_map.get(rg.get('resource_id'), '—'))
        seg_rules = fmt_filter_rules(seg.get('rules', []), col_map)

        sc2 = rg.get('segmentation_config') or {}
        crit_id = sc2.get('criteria_column_id')
        crit_col = col_map.get(crit_id, f"ID:{crit_id}") if crit_id else "—"
        segments = [m.get('value', '') for m in sc2.get('segmentation_metadata', []) if m.get('value')]

        groups.append({
            'prefix':        prefix,
            'resource_name': resource_name,
            'group_name':    seg_name,
            'group_filters': seg_rules,
            'crit_col':      crit_col,
            'segments':      segments,
        })

    rule_sets = []
    for rs in sorted(adv.get('reconciliation_rule_sets', []), key=lambda x: x.get('position', 99)):
        rules_desc = []
        for rule in (rs.get('reconciliation_rules') or []):
            col_a = col_map.get(rule.get('column_a_id'), f"ID:{rule.get('column_a_id')}")
            col_b = col_map.get(rule.get('column_b_id'), f"ID:{rule.get('column_b_id')}")
            op = rule.get('operator', '=')
            tol = rule.get('tolerance', 0)
            tol_u = rule.get('tolerance_unit') or ''
            tol_s = f"  [tolerancia ±{tol} {tol_u}]" if tol else ""
            rules_desc.append(f"A.{col_a}  {op}  B.{col_b}{tol_s}")

        sweep = []
        for sw in (rs.get('sweep_sides') or []):
            p = sw.get('prefix_side', '?')
            isr = sw.get('input_sweep_resource') or {}
            meta_id = isr.get('segmentation_metadata_id')
            if meta_id:
                seg_val = meta_map.get(meta_id, f"ID:{meta_id}")
            else:
                seg_val = "(recurso completo sin segmentar)"
            sweep.append(f"Lado {p}: {seg_val}")

        rule_sets.append({
            'pos':        rs.get('position', 0),
            'name':       rs.get('name', ''),
            'cross_type': rs.get('cross_type', ''),
            'new_ver':    rs.get('is_new_version', False),
            'rules':      rules_desc,
            'sweep':      sweep,
        })

    return {'groups': groups, 'rule_sets': rule_sets}

def parse_segment_filters(segs, col_map):
    result = []
    for seg in (segs or []):
        rules = []
        for fset in (seg.get('segment_filter_sets') or []):
            for r in (fset.get('segment_filter_rules') or []):
                col_name = col_map.get(r.get('column_id'), f"ID:{r.get('column_id')}")
                rules.append(
                    f"{r.get('condition','')} [{col_name}] {r.get('operator','')} {r.get('value','')}".strip()
                )
        if rules:
            result.append({'name': seg.get('name', ''), 'rules': rules})
    return result

def parse_source_group(sg, col_map):
    if not sg:
        return [], []
    group_cols = [col_map.get(c.get('column_id'), f"ID:{c.get('column_id')}")
                  for c in sorted(sg.get('columns', []), key=lambda x: x.get('position', 0))]
    agg_vals = [(v.get('function', '?'), col_map.get(v.get('column_id'), f"ID:{v.get('column_id')}"))
                  for v in sorted(sg.get('values', []), key=lambda x: x.get('position', 0))]
    return group_cols, agg_vals

def limpiar_hoja(nombre, eid):
    clean = re.sub(r'[\\/*?:\[\]]', '', str(nombre))
    return (clean[:18] + "_" + str(eid))[:31]

def sort_key(r):
    return (RT_ORDER.get(r.get('resource_type', ''), 99), r.get('export_id', 0))

def build_relations(resources, nodes, res_map):
    all_ids = {r.get('export_id') for r in resources}
    rels = {r.get('export_id'): {"parents": [], "children": []} for r in resources}
    for n in nodes:
        t_id = n.get('target')
        s_val = n.get('source')
        if not (t_id and s_val):
            continue
        s_list = s_val if isinstance(s_val, list) else [s_val]
        for sid in s_list:
            ext_a = "" if sid in all_ids else " ↗"
            ext_b = "" if t_id in all_ids else " ↗"
            if t_id in rels:
                rels[t_id]["parents"].append(res_map.get(sid, str(sid)) + ext_a)
            if sid in rels:
                rels[sid]["children"].append(res_map.get(t_id, str(t_id)) + ext_b)
    return rels

# ══════════════════════════════════════════════════════════════════════════════
# GENERADOR EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def generar_excel(data, selected_ids):
    all_resources = data.get('resources', [])
    nodes = data.get('nodes', [])
    res_map, col_map, seg_map, meta_map = build_maps(data)

    seen = set()
    resources = []
    for r in all_resources:
        eid = r.get('export_id')
        if eid in selected_ids and eid not in seen:
            seen.add(eid)
            resources.append(r)
            
    resources.sort(key=sort_key)
    rels = build_relations(resources, nodes, res_map)
    map_hojas = {r.get('export_id'): limpiar_hoja(r.get('name', ''), r.get('export_id')) for r in resources}

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        wb = writer.book

        # ── ÍNDICE ─────────────────────────────────────────────────────────────
        ws = wb.create_sheet("📚 Índice", 0)
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:H1')
        c = ws.cell(1, 1, "SIMETRIK DOCUMENTATION PRO  ·  PeYa Finance Operations & Control")
        sc(c, bg=C["red"], bold=True, color=C["white"], size=13, ha='center', va='center', wrap=False)
        ws.row_dimensions[1].height = 32

        ws.merge_cells('A2:H2')
        c = ws.cell(2, 1, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Recursos documentados: {len(resources)}")
        sc(c, bg=C["dark"], color=C["white"], size=9, ha='center', va='center', wrap=False)
        ws.row_dimensions[2].height = 15
        
        ws.freeze_panes = "A5"

        idx_hdrs = ["#", "ID", "NOMBRE DEL RECURSO", "TIPO", "PROVIENE DE", "ALIMENTA A", "ENCADENADA", "LINK"]
        for i, h in enumerate(idx_hdrs, 1):
            hdr(ws.cell(4, i, h), h, bg=C["dark"])
        ws.row_dimensions[4].height = 20

        for row_n, res in enumerate(resources, 5):
            eid = res.get('export_id')
            rt = res.get('resource_type', '')
            recon = res.get('reconciliation') or {}
            adv = res.get('advanced_reconciliation') or {}
            chained = recon.get('is_chained', False) or adv.get('is_chained', False)
            bg = C["grey"] if row_n % 2 == 0 else C["white"]

            vals = [
                row_n - 4, 
                eid, 
                res.get('name', ''), 
                RT_LABEL.get(rt, rt),
                ", ".join(rels[eid]["parents"]) or "— origen",
                ", ".join(rels[eid]["children"]) or "— fin de flujo",
                "Sí" if chained else "No"
            ]
            
            for col_n, val in enumerate(vals, 1):
                c = ws.cell(row_n, col_n, val)
                sc(c, bg=bg, size=9, va='center', wrap=False)
                if col_n == 4:
                    c.font = Font(name='Arial', bold=True, size=9, color=RT_COLOR.get(rt, C["dark"]))
                c.border = mk_border()

            lnk = ws.cell(row_n, 8, "Ver →")
            lnk.hyperlink = f"#'{map_hojas[eid]}'!A1"
            lnk.font = Font(name='Arial', color="0D47A1", underline="single", size=9)
            lnk.border = mk_border()
            ws.row_dimensions[row_n].height = 15

        for col_n, w in enumerate([6,11,44,26,36,36,12,8], 1):
            ws.column_dimensions[chr(64+col_n)].width = w
            
        ws.auto_filter.ref = f"A4:H{len(resources)+4}"

        # ── HOJAS DE DETALLE ───────────────────────────────────────────────────
        for res in resources:
            eid = res.get('export_id')
            rt = res.get('resource_type', '')
            name = res.get('name', '')
            tc = RT_COLOR.get(rt, C["dark"])
            COLS = 5

            ws = wb.create_sheet(map_hojas[eid])
            ws.sheet_view.showGridLines = False

            row = 1
            ws.merge_cells(f'A{row}:E{row}')
            c = ws.cell(row, 1, RT_LABEL.get(rt, '') + "  ·  " + name)
            sc(c, bg=tc, bold=True, color=C["white"], size=12, ha='left', va='center', wrap=False)
            ws.row_dimensions[row].height = 30
            row += 1

            row = meta_row(ws, row, "ID Recurso", eid, cols=COLS)
            row = meta_row(ws, row, "Tipo", RT_LABEL.get(rt, rt), cols=COLS)
            row = meta_row(ws, row, "Proviene de", ", ".join(rels[eid]["parents"]) or "Origen", cols=COLS)
            row = meta_row(ws, row, "Alimenta a", ", ".join(rels[eid]["children"]) or "Fin de flujo", cols=COLS)
            
            ws.freeze_panes = f"A{row+1}"
            row += 1

            # ── CONCILIACIÓN ESTÁNDAR ──────────────────────────────────────────
            std = parse_std_reconciliation(res.get('reconciliation'), res_map, col_map, seg_map)
            if std:
                row = section_title(ws, row, "⚖️  REGLAS DE CONCILIACIÓN ESTÁNDAR", bg=C["red"], cols=COLS)

                row = section_title(ws, row, "  GRUPOS CONCILIABLES ACTIVOS", bg=C["rose"], cols=COLS)
                for col_n, h in enumerate(["LADO", "RECURSO", "GRUPO CONCILIABLE (ACTIVO)", "FILTROS DEL GRUPO"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg=C["rose"])
                ws.merge_cells(f'D{row}:E{row}')
                ws.row_dimensions[row].height = 18
                row += 1

                for i, side in enumerate(std['sides']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    trig = "  [TRIGGER]" if side['is_trigger'] else ""
                    
                    c1 = ws.cell(row, 1, side['prefix'] + trig)
                    c2 = ws.cell(row, 2, side['resource_name'])
                    c3 = ws.cell(row, 3, side['group_name'])
                    ws.merge_cells(f'D{row}:E{row}')
                    c4 = ws.cell(row, 4, side['group_filters'])
                    
                    sc(c1, bg=bg, size=9, va='top', wrap=True, ha='center')
                    sc(c2, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c3, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c4, bg=bg, size=9, va='top', wrap=True, ha='left')
                    
                    ws.row_dimensions[row].height = row_height(side['group_filters'].count('\n') + 1)
                    row += 1
                row += 1

                row = meta_row(ws, row, "Conciliación encadenada", "Sí" if std['is_chained'] else "No", cols=COLS)
                row += 1

                row = section_title(ws, row, "  RULE SETS DE MATCHING", bg="C62828", cols=COLS)
                for col_n, h in enumerate(["POS.", "NOMBRE DEL RULE SET", "REGLAS  (A vs B)"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg="C62828")
                ws.merge_cells(f'C{row}:E{row}')
                ws.row_dimensions[row].height = 18
                row += 1

                for i, rs in enumerate(std['rule_sets']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    
                    c1 = ws.cell(row, 1, rs['pos'])
                    c2 = ws.cell(row, 2, rs['name'])
                    ws.merge_cells(f'C{row}:E{row}')
                    c3 = ws.cell(row, 3, "\n".join(rs['rules']))
                    
                    sc(c1, bg=bg, size=9, va='top', wrap=True, ha='center')
                    sc(c2, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c3, bg=bg, size=9, va='top', wrap=True, ha='left')
                    
                    ws.row_dimensions[row].height = row_height(len(rs['rules']))
                    row += 1
                row += 1

            # ── CONCILIACIÓN AVANZADA ──────────────────────────────────────────
            adv_parsed = parse_adv_reconciliation(res.get('advanced_reconciliation'), res_map, col_map, seg_map, meta_map)
            if adv_parsed:
                row = section_title(ws, row, "🔬  REGLAS DE CONCILIACIÓN AVANZADA", bg=C["purple"], cols=COLS)

                row = section_title(ws, row, "  GRUPOS CONCILIABLES Y SEGMENTOS INTERNOS", bg="6A1B9A", cols=COLS)
                for col_n, h in enumerate(["LADO", "RECURSO", "GRUPO CONCILIABLE", "FILTROS DEL GRUPO", "SEGMENTOS INTERNOS"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg="6A1B9A")
                ws.row_dimensions[row].height = 18
                row += 1

                for i, g in enumerate(adv_parsed['groups']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    segs_txt = "\n".join(g['segments']) if g['segments'] else "(sin segmentación interna)"
                    n_lines = max(g['group_filters'].count('\n') + 1, len(g['segments']) if g['segments'] else 1)
                    
                    c1 = ws.cell(row, 1, g['prefix'])
                    c2 = ws.cell(row, 2, g['resource_name'])
                    c3 = ws.cell(row, 3, g['group_name'])
                    c4 = ws.cell(row, 4, g['group_filters'])
                    c5 = ws.cell(row, 5, segs_txt)
                    
                    sc(c1, bg=bg, size=9, va='top', wrap=True, ha='center')
                    sc(c2, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c3, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c4, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c5, bg=bg, size=9, va='top', wrap=True, ha='left')
                    
                    ws.row_dimensions[row].height = row_height(n_lines)
                    row += 1
                row += 1

                row = section_title(ws, row, "  RULE SETS (SEGMENTO A vs SEGMENTO B)", bg="4A148C", cols=COLS)
                for col_n, h in enumerate(["POS.", "NOMBRE / TIPO", "REGLAS  (A vs B)", "SEGMENTO LADO A", "SEGMENTO LADO B"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg="4A148C")
                ws.row_dimensions[row].height = 18
                row += 1

                for i, rs in enumerate(adv_parsed['rule_sets']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    name_txt = rs['name']
                    if rs['cross_type']:
                        name_txt += "\n[" + rs['cross_type'] + "]"
                    if rs['new_ver']:
                        name_txt += "  ✦ new version"

                    seg_a = next((s.replace("Lado A: ", "") for s in rs['sweep'] if s.startswith("Lado A")), "—")
                    seg_b = next((s.replace("Lado B: ", "") for s in rs['sweep'] if s.startswith("Lado B")), "—")

                    c1 = ws.cell(row, 1, rs['pos'])
                    c2 = ws.cell(row, 2, name_txt)
                    c3 = ws.cell(row, 3, "\n".join(rs['rules']))
                    c4 = ws.cell(row, 4, seg_a)
                    c5 = ws.cell(row, 5, seg_b)
                    
                    sc(c1, bg=bg, size=9, va='top', wrap=True, ha='center')
                    sc(c2, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c3, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c4, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c5, bg=bg, size=9, va='top', wrap=True, ha='left')
                    
                    n_lines = max(len(rs['rules']), 1)
                    ws.row_dimensions[row].height = row_height(n_lines)
                    row += 1
                row += 1

            # ── SOURCE GROUP ──────────────────────────────────────────────────
            sg = res.get('source_group')
            if sg:
                row = section_title(ws, row, "📊  CONFIGURACIÓN DE AGRUPACIÓN (GROUP BY)", bg=C["amber"], cols=COLS)
                group_cols, agg_vals = parse_source_group(sg, col_map)
                
                row = meta_row(ws, row, "GROUP BY (dimensiones)", " | ".join(group_cols) or "—", cols=COLS, bg_val="FFF3E0")
                
                agg_str = "  |  ".join(f"{fn}( {col} )" for fn, col in agg_vals)
                row = meta_row(ws, row, "Agregaciones (métricas)", agg_str or "—", cols=COLS, bg_val="FFF3E0")
                
                row = meta_row(ws, row, "Acumulativo", "Sí" if sg.get('is_accumulative') else "No", cols=COLS)
                row += 1

            # ── SOURCE UNION ──────────────────────────────────────────────────
            su = res.get('source_union')
            if su:
                row = section_title(ws, row, "🔗  CONFIGURACIÓN DE UNIÓN DE FUENTES", bg=C["teal"], cols=COLS)
                for us in (su.get('union_segments') or []):
                    label = "TRIGGER" if us.get('is_trigger') else "No trigger"
                    ttype = us.get('trigger_type') or ''
                    val_str = f"{label}  {'· ' + ttype if ttype else ''}"
                    row = meta_row(ws, row, f"Segmento ID {us.get('segment_id', '')}", val_str, cols=COLS, bg_val="E0F2F1")
                row += 1

            # ── GRUPOS CONCILIABLES DEL RECURSO ──────────────────────────────
            segs_all = parse_segment_filters(res.get('segments', []), col_map)
            if segs_all:
                row = section_title(ws, row, "🔍  GRUPOS CONCILIABLES DEL RECURSO", bg=C["slate"], cols=COLS)
                for col_n, h in enumerate(["NOMBRE DEL GRUPO", "FILTROS APLICADOS"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg=C["slate"])
                ws.merge_cells(f'B{row}:E{row}')
                ws.row_dimensions[row].height = 18
                row += 1
                
                for i, seg in enumerate(segs_all):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    c1 = ws.cell(row, 1, seg['name'])
                    ws.merge_cells(f'B{row}:E{row}')
                    c2 = ws.cell(row, 2, "\n".join(seg['rules']))
                    
                    sc(c1, bg=bg, size=9, va='top', wrap=True)
                    sc(c2, bg=bg, size=9, va='top', wrap=True)
                    
                    ws.row_dimensions[row].height = row_height(len(seg['rules']))
                    row += 1
                row += 1

            # ── COLUMNAS ──────────────────────────────────────────────────────
            columns = sorted(res.get('columns') or [], key=lambda x: x.get('position', 0))
            if columns:
                row = section_title(ws, row, "📋  CONFIGURACIÓN DE COLUMNAS", bg=C["blue"], cols=COLS)
                
                start_filter_row = row
                for col_n, h in enumerate(["LABEL / NOMBRE", "TIPO DATO", "TIPO COL.", "LÓGICA · FÓRMULA · BUSCAR V"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg=C["blue"])
                ws.merge_cells(f'D{row}:E{row}')
                ws.row_dimensions[row].height = 18
                ws.auto_filter.ref = f"A{start_filter_row}:E{start_filter_row + len(columns)}"
                row += 1
                
                for i, col in enumerate(columns):
                    label = col.get('label') or col.get('name', '')
                    dtype = col.get('data_format', '')
                    col_type = (col.get('column_type') or '').replace('_', ' ').upper()
                    logic = parse_transformation_logic(col, res_map, col_map)
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    
                    c1 = ws.cell(row, 1, label)
                    c2 = ws.cell(row, 2, dtype)
                    c3 = ws.cell(row, 3, col_type)
                    ws.merge_cells(f'D{row}:E{row}')
                    c4 = ws.cell(row, 4, logic)
                    
                    sc(c1, bg=bg, size=9, va='top', wrap=True, ha='left')
                    sc(c2, bg=bg, size=9, va='top', wrap=True, ha='center')
                    sc(c3, bg=bg, size=9, va='top', wrap=True, ha='center')
                    sc(c4, bg=bg, size=9, va='top', wrap=True, ha='left')
                    
                    ws.row_dimensions[row].height = row_height(logic.count('\n') + 1)
                    row += 1

            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 28
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 22

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    output.seek(0)
    return output

# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style='background:linear-gradient(135deg,#EA0050 0%,#B0003A 100%);
    padding:28px 36px;border-radius:16px;
    box-shadow:0 6px 24px rgba(234,0,80,0.25);margin-bottom:24px'>
    <h1 style='color:white;margin:0;font-family:Arial,sans-serif;
        font-size:2.2rem;letter-spacing:-0.5px;font-weight:700'>
        📊 Simetrik Docs Pro
    </h1>
    <p style='color:rgba(255,255,255,0.88);margin:8px 0 0;
        font-size:1.05rem;font-family:Arial'>
        PeYa Finance Operations &amp; Control &nbsp;·&nbsp;
        Generador Automático de Documentación
    </p>
</div>""", unsafe_allow_html=True)

up = st.file_uploader("📂 Arrastrá o subí el JSON exportado de tu flujo de Simetrik", type=['json'])

if not up:
    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    c1.info("**1. En Simetrik**\n\nVe a tu Flujo de trabajo y haz clic en ⚙️ **Configuración**.")
    c2.info("**2. Exportar**\n\nBusca la opción **Exportar JSON** en el menú de la derecha.")
    c3.success("**3. Generar**\n\nSube el archivo aquí y obtén tu Excel corporativo al instante.")
    st.stop()

try:
    data = json.load(up)
    all_resources = data.get('resources', [])
    nodes = data.get('nodes', [])

    seen_load = set()
    resources_unique = []
    for r in all_resources:
        eid = r.get('export_id')
        if eid and eid not in seen_load:
            seen_load.add(eid)
            resources_unique.append(r)

    resources_unique.sort(key=sort_key)
    res_map, col_map, seg_map, meta_map = build_maps(data)
    rels_all = build_relations(resources_unique, nodes, res_map)

except Exception as e:
    st.error(f"Error al analizar la estructura del JSON: {e}")
    st.stop()

if 'sel' not in st.session_state:
    st.session_state.sel = {r.get('export_id'): True for r in resources_unique}

# --- DASHBOARD DE MÉTRICAS ---
st.markdown("### 📈 Resumen del Flujo")
m1, m2, m3, m4 = st.columns(4)
m1.metric("Total Recursos", len(resources_unique))
m2.metric("Fuentes Originales", len([r for r in resources_unique if r.get('resource_type') == 'native']))
m3.metric("Reglas de Conciliación", len([r for r in resources_unique if 'reconciliation' in r.get('resource_type', '')]))
m4.metric("Nodos de Conexión", len(nodes))
st.markdown("---")

# --- SIDEBAR ---
with st.sidebar:
    st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/e/e0/PedidosYa_logo.svg/512px-PedidosYa_logo.svg.png", width=150)
    st.markdown("### ⚙️ Configuración")
    
    all_types = sorted({r.get('resource_type', '') for r in resources_unique}, key=lambda x: RT_ORDER.get(x, 99))
    filtro_tipo = st.multiselect(
        "Filtrar por tipo:", 
        options=all_types, 
        format_func=lambda x: RT_LABEL.get(x, x), 
        default=all_types
    )
    
    resources_visible = [r for r in resources_unique if r.get('resource_type', '') in filtro_tipo]
    
    st.markdown("#### Selección Rápida")
    c_btn1, c_btn2 = st.columns(2)
    if c_btn1.button("✅ Todos", use_container_width=True):
        for r in resources_visible:
            st.session_state.sel[r.get('export_id')] = True
    if c_btn2.button("⬜ Ninguno", use_container_width=True):
        for r in resources_visible:
            st.session_state.sel[r.get('export_id')] = False

# --- ÁREA PRINCIPAL ---
st.markdown("### 1️⃣  Seleccioná los recursos a documentar")

tipo_groups = {}
for r in resources_visible:
    tipo_groups.setdefault(r.get('resource_type', ''), []).append(r)

selected_ids = set()

for rt in sorted(tipo_groups.keys(), key=lambda x: RT_ORDER.get(x, 99)):
    grupo = tipo_groups[rt]
    label_tipo = RT_LABEL.get(rt, rt)
    
    with st.expander(f"{label_tipo} ({len(grupo)} recursos)", expanded=False):
        for r in grupo:
            eid = r.get('export_id')
            ca, cb, cc, cd = st.columns([0.5, 4, 3, 3])
            
            checked = ca.checkbox("", value=st.session_state.sel.get(eid, True), key=f"chk_{eid}")
            st.session_state.sel[eid] = checked
            
            cb.markdown(f"**{r.get('name','')}** <span style='color:gray;font-size:0.8em'>`ID:{eid}`</span>", unsafe_allow_html=True)
            cc.caption(f"⬅️ {', '.join(rels_all[eid]['parents']) or '—'}")
            cd.caption(f"➡️ {', '.join(rels_all[eid]['children']) or '—'}")
            
            if checked:
                selected_ids.add(eid)

st.markdown("---")

n_sel = len(selected_ids)
if not selected_ids:
    st.warning("⚠️ Seleccioná al menos un recurso en la lista superior.")
    st.stop()

if st.button(f"🚀  GENERAR EXCEL DE DOCUMENTACIÓN ({n_sel} recursos)", type="primary", use_container_width=True):
    with st.status("Preparando documento corporativo...", expanded=True) as status:
        try:
            st.write("Resolviendo reglas de Duplicados, Particiones y Búsquedas V...")
            excel_bytes = generar_excel(data, selected_ids)
            status.update(label=f"✅ ¡Completado! Documentación generada con éxito.", state="complete", expanded=False)
            
            st.download_button(
                label="📥  DESCARGAR REPORTE EXCEL",
                data=excel_bytes,
                file_name=f"DOC_Simetrik_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            st.balloons()
        except Exception as e:
            status.update(label="❌ Error durante la generación", state="error")
            st.error(str(e))
